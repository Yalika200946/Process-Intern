"""Read design-only HX references from the plant workbook with explicit OCR quality."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl


SHEET_GROUP = {
    "3E-101A": "E101AB", "3E-101B": "E101AB", "3E-101C": "E101CD", "3E-101D": "E101CD",
    "3E-101E": "E101EF", "3E-101F": "E101EF", "3E-102": "E102", "3E-103A": "E103AB",
    "3E-103B": "E103AB", "3E-104": "E104", "3E-105A": "E105AB", "3E-105B": "E105AB",
    "3E-106A": "E106AB", "3E-106B": "E106AB", "3E-107A": "E107AB", "3E-107B": "E107AB",
}


def _number(value):
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str) or "OCR unclear" in value or "—" in value:
        return None
    match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", value)
    return float(match.group(0).replace(",", "")) if match else None


def _labeled_value(rows, label: str):
    for row in rows:
        for index, cell in enumerate(row[:-1]):
            if isinstance(cell, str) and label.lower() in cell.lower():
                for candidate in row[index + 1:]:
                    value = _number(candidate)
                    if value is not None:
                        return value
    return None


def read_datasheet_diagnostics(path: Path) -> dict[str, dict]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    grouped: dict[str, list[dict]] = {}
    for sheet, hx in SHEET_GROUP.items():
        if sheet not in workbook.sheetnames:
            continue
        rows = list(workbook[sheet].iter_rows(values_only=True))
        area = _labeled_value(rows, "Surf./Train") or _labeled_value(rows, "Surf./Shell")
        clean_u = _labeled_value(rows, "Transfer Rate")
        # Prefer the explicit Clean row when both service and clean transfer rates exist.
        for row in rows:
            if row[0] and "Transfer Rate" in str(row[0]) and "Clean" in str(row[0]):
                clean_u = next((_number(v) for v in row[1:] if _number(v) is not None), None)
        service_u = next((next((_number(v) for v in row[1:] if _number(v) is not None), None)
                          for row in rows if row[0] and "Transfer Rate" in str(row[0]) and "Service" in str(row[0])), None)
        lmtd = _labeled_value(rows, "MTD (Corrected)")
        duty = _labeled_value(rows, "Heat Exchanged")
        quality = "USABLE" if area and clean_u and lmtd else "PARTIAL_OCR_OR_MISSING"
        grouped.setdefault(hx, []).append({"sheet": sheet, "area_m2": area, "u_clean_w_m2k": clean_u,
                                           "u_service_w_m2k": service_u, "lmtd_c": lmtd,
                                           "design_duty_kw": duty, "quality": quality})
    output = {}
    for hx, sheets in grouped.items():
        usable = [s for s in sheets if s["quality"] == "USABLE"]
        output[hx] = {"sheets": sheets, "coverage": "AVAILABLE" if usable else "PARTIAL",
                      "ua_clean_kw_k": round(sum(s["area_m2"] * s["u_clean_w_m2k"] / 1000 for s in usable), 4) if usable else None,
                      "diagnostic_clean_duty_kw": round(sum(s["area_m2"] * s["u_clean_w_m2k"] * s["lmtd_c"] / 1000 for s in usable), 2) if usable else None,
                      "approval_status": "CANDIDATE", "basis": "design datasheet; not current plant UA"}
    return output
