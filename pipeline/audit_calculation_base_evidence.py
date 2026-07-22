"""Read-only inventory of Calculation base Example.xlsx area evidence."""
from __future__ import annotations

import json
from pathlib import Path
import sys

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path: sys.path.insert(0, str(ROOT))
from src.calculations.overall_heat_transfer import calculate_u_from_ua

SOURCE = Path(r"C:\Desktop\Bangchak Internship 2026\Data\Calculation base Example.xlsx")
OUT = ROOT / "reports/tables/mvp_real_data/dependency_closure"
CANONICAL = {"3E-101AB":"E101AB","3E-101C":"E101CD","3E-101D":"E101CD","3E-101E":"E101EF",
             "3E-101F":"E101EF","3E-102":"E102","3E-103AB":"E103AB","3E-104":"E104",
             "3E-105AB":"E105AB","3E-106AB":"E106AB","3E-107AB":"E107AB"}


def inventory_workbook(path: Path = SOURCE) -> pd.DataFrame:
    formula = openpyxl.load_workbook(path, read_only=True, data_only=False)
    values = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows=[]
    for sheet in values.sheetnames:
        ws, wf = values[sheet], formula[sheet]
        for row_i in range(1, ws.max_row + 1):
            source_name = ws.cell(row_i,18).value
            if not isinstance(source_name,str) or source_name not in CANONICAL: continue
            area=ws.cell(row_i,38).value; shells=ws.cell(row_i,39).value; u_service=ws.cell(row_i,37).value
            purpose=ws.cell(row_i,19).value; remarks=ws.cell(row_i,42).value
            has_area=isinstance(area,(int,float)) and area>0
            ambiguity="Example workbook; no approval/revision; active-shell basis not independently verified."
            if has_area and shells: ambiguity += f" Header says Area/Train with {shells} shells, but per-shell versus active-train basis remains unapproved."
            rows.append({"file_name":path.name,"file_path":str(path),"file_type":path.suffix,
                         "worksheet_or_section":f"{sheet}!row {row_i}","source_exchanger_name":source_name,
                         "exchanger_id":CANONICAL[source_name],"mapping_method":"EXACT_EXAMPLE_LABEL",
                         "shell_coverage":f"reported shells={shells}" if shells else "UNKNOWN",
                         "mapping_confidence":"MEDIUM" if has_area else "LOW",
                         "calculation_purpose":purpose,"area_value":area if has_area else None,
                         "area_unit":"m2" if has_area else None,"area_basis":"Area A/Train header; example basis" if has_area else "UNKNOWN",
                         "source_or_reference":remarks,"design_or_example":"EXAMPLE_CALCULATION_WORKBOOK",
                         "date_or_revision":path.stat().st_mtime,"u_service_value":u_service,
                         "u_service_unit":"W/m2/K" if isinstance(u_service,(int,float)) else None,
                         "usable_status":"EXAMPLE_AREA_ONLY" if has_area else "NOT_USABLE",
                         "ambiguity":ambiguity,"notes":"No source file was modified."})
    return pd.DataFrame(rows)


def ua_u_audit(inventory: pd.DataFrame) -> pd.DataFrame:
    rows=[]
    for _,r in inventory.iterrows():
        probe=calculate_u_from_ua(ua_value=100.0,ua_unit="kW/K",area_value=r.area_value or 1,
                                  area_unit=r.area_unit or "m2",area_status=r.usable_status,
                                  f_status="ASSUMPTION_F1",shell_basis_matches=False)
        rows.append({"hx_id":r.exchanger_id,"ua_value":None,"ua_unit":"kW/K",
                     "area_value":r.area_value,"area_unit":r.area_unit,"area_basis":r.area_basis,
                     "area_source":r.file_path,"area_status":r.usable_status,"F_value":1.0,
                     "F_source":"config/mvp_real_data_pilot.json","F_status":"ASSUMPTION",
                     "u_value":probe["u_value"],"u_unit":"W/m2/K","u_status":probe["u_status"],
                     "existing_metric_audit":"Preserve apparent UA(F=1); do not publish plant U from this evidence."})
    return pd.DataFrame(rows)


def main():
    OUT.mkdir(parents=True,exist_ok=True)
    inv=inventory_workbook();inv.to_csv(OUT/"calculation_base_inventory.csv",index=False)
    audit=ua_u_audit(inv);audit.to_csv(OUT/"calculation_base_ua_u_audit.csv",index=False)
    mappings=inv[["source_exchanger_name","exchanger_id","mapping_method","shell_coverage","mapping_confidence","source_or_reference","ambiguity","usable_status"]]
    mappings.to_csv(OUT/"calculation_base_hx_mapping.csv",index=False)
    summary={"files_inspected":1,"worksheets_inspected":int(inv.worksheet_or_section.str.split("!").str[0].nunique()),
             "hx_sections_found":len(inv),"area_values_found":int(inv.area_value.notna().sum()),
             "verified_areas":int(inv.usable_status.isin(["VERIFIED_DESIGN_AREA","VERIFIED_CALCULATION_AREA"]).sum()),
             "example_only_areas":int(inv.usable_status.eq("EXAMPLE_AREA_ONLY").sum()),
             "hx_newly_eligible_for_u":[],"conclusion":"No example area is eligible for current plant U."}
    (OUT/"calculation_base_summary.json").write_text(json.dumps(summary,indent=2),encoding="utf-8")
    print(inv[["exchanger_id","area_value","area_unit","area_basis","usable_status","ambiguity"]].to_string(index=False))

if __name__=="__main__":main()
